from __future__ import annotations

import json
import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path.cwd()
WORKSPACE_ROOT = PROJECT_ROOT.parent
SOURCE_XLSX = next(
    path for path in WORKSPACE_ROOT.iterdir()
    if path.suffix.lower() == ".xlsx" and "v1.0" in path.name
)
OUTPUT_JS = PROJECT_ROOT / "story-data-v1.js"

STORY_SHEETS = [
    "EP1 첫만남",
    "EP2 동아리 OT",
    "EP3 동아리 MT",
    "EP4-A 넷이",
    "EP4-B 둘이",
    "EP5 미팅사건",
    "EP6 정기공연",
]

MINIGAME_MOVES = {
    "EP1 첫만남": {
        "type": "move",
        "next": "MINIGAME",
        "minigame": "brickBreaker",
        "after": "EP2 동아리 OT",
        "options": {
            "maxTurns": 5,
            "tutorial": [
                "첫 벽돌깨기는 5턴만 진행돼. 마우스로 각도를 정하고 클릭하면 공이 나가.",
                "블럭 숫자는 남은 내구도야. 목표는 끝날 때 도파민을 적당히 남기는 거야.",
                "빨간 블럭은 맞힐 때마다 도파민이 +1씩 늘어나고, 파란 블럭은 부숴지면 도파민이 8 감소해.",
                "블럭이 아래까지 내려오면 바로 끝나지만, 그 순간의 도파민이 결과로 돌아가.",
            ],
        },
    },
    "EP2 동아리 OT": {
        "type": "move",
        "next": "MINIGAME",
        "minigame": "sideShooter",
        "after": "EP3 동아리 MT",
        "options": {
            "difficulty": 0,
            "durationSeconds": 45,
            "tutorial": [
                "슈팅은 45초만 버티면 돼. 마우스를 움직이면 캐릭터도 따라 움직여.",
                "노란 P 캡슐을 먹으면 우측 위 파워 칸이 한 칸 올라가.",
                "원하는 칸에서 우클릭하면 그 기술을 쓸 수 있어.",
                "기술을 쓰면 화면에 빛 효과가 떠서 어떤 행동이 발동됐는지 바로 확인할 수 있어.",
                "흡수 기술은 도파민을 크게 낮춰. 속도 증가는 더 빠르게 움직이고 도파민도 조금 더 빨리 낮춰.",
                "적들에게 맞으면 체력이 줄고, 도파민이 증가해.",
                "체력은 3이고, 체력이 0이되면 그때의 도파민으로 게임이 끝나.",
            ],
        },
    },
    "EP4-A 넷이": {
        "type": "move",
        "next": "MINIGAME",
        "minigame": "sideShooter",
        "after": "EP5 미팅사건",
        "options": {"difficulty": 1, "durationSeconds": 45},
    },
    "EP4-B 둘이": {
        "type": "move",
        "next": "MINIGAME",
        "minigame": "sideShooter",
        "after": "EP5 미팅사건",
        "options": {"difficulty": 1, "durationSeconds": 45},
    },
    "EP5 미팅사건": {
        "type": "move",
        "next": "MINIGAME",
        "minigame": "brickBreaker",
        "after": "EP6 정기공연",
        "options": {"maxTurns": 6},
    },
}

EP3_BRANCH_MOVE = {
    "type": "move",
    "next": "MINIGAME",
    "minigame": "brickBreaker",
    "options": {"maxTurns": 6},
}

SOUND_EFFECTS = {
    "띠링 소리": "ep1Ding",
    "자동문 열리는 소리": "ep1Ding",
    "바코드 찍는 소리": "ep1Barcode",
    "의자 끄는 소리": "chair",
    "왁자지껄한 소리": "crowd",
    "무서운 소리": "horror",
    "문 여는 효과음": "door",
    "달리는 소리": "running",
    "드라이버 소리": "screwdriver",
    "박수 소리": "crowd",
}


@dataclass
class Row:
    index: int
    selection: str
    speaker: str
    text: str
    dopamine: int | None
    affection: int | None
    characters: str
    emotion: str
    background: str
    sound: str
    note: str
    is_gray: bool
    is_yellow: bool


class EpisodeBuilder:
    def __init__(self, episode_id: str):
        self.episode_id = episode_id
        self.nodes: list[dict] = []
        self.next_id = 1
        self.current_background = ""
        self.visible_characters: dict[str, str] = {}
        self.last_prompt = "선택"
        self.pending_transition: dict | None = None
        self.add({"type": "scene reset"})

    def add(self, node: dict) -> dict:
        node["id"] = self.next_id
        self.next_id += 1
        self.nodes.append(node)
        return node

    def ensure_background(self, name: str, transition: dict | None = None) -> None:
        if not name or name == self.current_background:
            return

        node = {"type": "background", "name": name}
        if transition:
            node["transition"] = transition
        self.add(node)
        self.current_background = name
        self.clear_characters()
        if is_cg_background(name):
            return

    def clear_characters(self) -> None:
        if not self.visible_characters:
            return

        self.add({"type": "clear characters"})
        self.visible_characters = {}

    def add_row_characters(self, row: Row) -> None:
        if is_cg_background(row.background or self.current_background):
            return

        condition = parse_condition(row.selection)
        for name in parse_characters(row.characters):
            emotion = row.emotion or "Normal"
            if self.visible_characters.get(name) == emotion:
                continue

            node = {
                "type": "character in",
                "name": name,
                "emotion": emotion,
            }
            if condition:
                node["condition"] = condition
            self.add(node)
            self.visible_characters[name] = emotion

    def add_sounds(self, raw_sound: str) -> None:
        for sound in parse_sounds(raw_sound, self.episode_id):
            self.add(sound)

    def add_dialogue(self, row: Row) -> None:
        if not row.text:
            return
        if is_editor_note_text(row.text):
            return

        speaker = normalize_speaker(row.speaker, row.text)
        node = {
            "type": "dialogue",
            "speaker": speaker,
            "text": normalize_text(row.text, speaker),
        }
        effects = parse_effects(row)
        if effects:
            node["effects"] = effects
        condition = parse_condition(row.selection)
        if condition:
            node["condition"] = condition
        self.add(node)
        self.last_prompt = node["text"]


def text_value(value) -> str:
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def parse_number(value) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(str(value).replace("+", "").strip()))
    except ValueError:
        return None


def fill_rgb(cell) -> str:
    color = cell.fill.fgColor
    if color.type == "rgb":
        return (color.rgb or "").upper()
    return ""


def is_gray_row(ws, row_index: int) -> bool:
    fills = [fill_rgb(ws.cell(row_index, col)) for col in range(1, 11)]
    return any(fill.endswith("999999") for fill in fills)


def is_yellow_row(ws, row_index: int) -> bool:
    fills = [fill_rgb(ws.cell(row_index, col)) for col in range(1, 11)]
    return any(fill.endswith("FFFF00") for fill in fills)


def read_rows(ws) -> list[Row]:
    rows: list[Row] = []
    for idx in range(3, ws.max_row + 1):
        values = [text_value(ws.cell(idx, col).value) for col in range(1, 11)]
        gray = is_gray_row(ws, idx)
        if not any(values) and not gray:
            continue

        rows.append(Row(
            index=idx,
            selection=values[0],
            speaker=values[1],
            text=values[2],
            dopamine=parse_number(values[3]),
            affection=parse_number(values[4]),
            characters=values[5],
            emotion=values[6],
            background=normalize_background(values[7]),
            sound=values[8],
            note=values[9],
            is_gray=gray,
            is_yellow=is_yellow_row(ws, idx),
        ))
    return rows


def normalize_background(name: str) -> str:
    return name.strip()


def normalize_text(text: str, speaker: str = "") -> str:
    normalized = text.strip().replace("000", "00").replace("OO", "00")
    normalized = re.sub(r"\n+", "\n", normalized)
    if speaker == "독백" and not normalized.startswith("("):
        return f"({normalized})"
    return normalized


def normalize_speaker(speaker: str, text: str = "") -> str:
    if speaker:
        return speaker
    return "지시문" if text else ""


def is_cg_background(name: str) -> bool:
    if not name:
        return False
    if "(CG)" in name:
        return True
    return name in {
        "소주병을 향해 뻗는 수진의 손",
        "소주병을 향해 뻗는 수진의 손을 잡는 건호의 손",
        "예쁜 파스타",
        "수진이 우는 모습",
        "토닥이는 장면",
        "휴지를 건네는 장면",
        "무대막이 내리는 장면",
    }


def parse_characters(raw_characters: str) -> list[str]:
    if not raw_characters:
        return []

    names = re.split(r"[,/·&\s]+", raw_characters.strip())
    return [
        name for name in names
        if name and name != "주인공"
    ]


def parse_effects(row: Row) -> dict:
    effects = {}
    if row.dopamine is not None:
        effects["dopamine"] = row.dopamine
    if row.affection is not None:
        effects["affection"] = row.affection
    return effects


def parse_condition(selection: str) -> dict | None:
    state = get_state(selection)
    return {"dopamineState": state} if state else None


def get_state(selection: str) -> str:
    match = re.match(r"^(LOW|OPT|HIGH)\b", selection.strip(), re.IGNORECASE)
    return match.group(1).upper() if match else ""


def is_choice_row(row: Row) -> bool:
    selection = row.selection.replace(" ", "")
    if selection.startswith("선택"):
        return True
    return bool(re.match(r"^(LOW|OPT|HIGH)\d+$", selection, re.IGNORECASE))


def is_follow_row(row: Row) -> bool:
    return row.selection.strip() == "-"


def is_marker_row(row: Row) -> bool:
    return row.selection.startswith("<") and row.selection.endswith(">")


def is_editor_note_text(text: str) -> bool:
    return bool(re.match(r"^<(해피|배드|베드)엔딩,.*>$", text.strip()))


def note_has_fade(row: Row) -> bool:
    return "fade" in row.note.lower()


def outdoor_transition(name: str) -> bool:
    return any(token in name for token in ["외부", "길거리", "야외", "밤, 대학가", "나와 걸어가는"])


def make_transition(name: str) -> dict:
    if outdoor_transition(name):
        return {
            "type": "fadeSlide",
            "duration": 1200,
            "slideDuration": 850,
            "direction": "left",
        }
    return {"type": "fadeBlack"}


def bgm_for_episode(episode_id: str) -> str:
    if episode_id.startswith("EP2") or episode_id.startswith("EP5"):
        return "ep25"
    if episode_id.startswith("EP3") or episode_id.startswith("EP6"):
        return "ep36"
    return "ep14"


def parse_sounds(raw_sound: str, episode_id: str) -> list[dict]:
    if not raw_sound:
        return []

    result = []
    parts = [part.strip() for part in raw_sound.split(",") if part.strip()]
    for part in parts:
        if part == "BGM 시작":
            result.append({"type": "sound", "soundType": "bgm", "action": "play", "name": bgm_for_episode(episode_id)})
        elif part == "공포 BGM 시작":
            result.append({"type": "sound", "soundType": "bgm", "action": "play", "name": "escapeRoom"})
        elif part in SOUND_EFFECTS:
            result.append({"type": "sound", "soundType": "effect", "action": "play", "name": SOUND_EFFECTS[part]})
    return result


def make_follow_line(row: Row) -> dict:
    speaker = normalize_speaker(row.speaker, row.text)
    line: dict = {}
    if row.background:
        line["background"] = row.background
        if is_cg_background(row.background):
            line["clearCharacters"] = True
    if note_has_fade(row):
        line["transition"] = make_transition(row.background)
    characters = parse_characters(row.characters)
    if characters and not is_cg_background(row.background):
        line["characters"] = [
            {"name": name, "emotion": row.emotion or "Normal"}
            for name in characters
        ]
    sounds = parse_sounds(row.sound, "")
    if sounds:
        line["sound"] = sounds[-1]
    if speaker:
        line["speaker"] = speaker
    if row.text:
        line["text"] = normalize_text(row.text, speaker)
    effects = parse_effects(row)
    if effects:
        line["effects"] = effects
    return line


def make_choice(row: Row) -> dict:
    choice = {
        "text": normalize_text(row.text, normalize_speaker(row.speaker, row.text)),
        "follow": [],
    }
    effects = parse_effects(row)
    if effects:
        choice["effects"] = effects
    condition = parse_condition(row.selection)
    if condition:
        choice["condition"] = condition
    sounds = parse_sounds(row.sound, "")
    if sounds:
        choice["sound"] = sounds[-1]
    if row.is_yellow:
        choice["disabledPreview"] = True
    return choice


def branch_target_from_note(note: str) -> str:
    if "4-A" in note:
        return "EP4-A 넷이"
    if "4-B" in note:
        return "EP4-B 둘이"
    return ""


def find_next_content_background(rows: list[Row], start_index: int) -> str:
    for row in rows[start_index:]:
        if row.is_gray:
            continue
        if row.background:
            return row.background
    return ""


def previous_background(rows: list[Row], row_index: int) -> str:
    for row in reversed(rows[:row_index]):
        if not row.is_gray and row.background:
            return row.background
    return ""


def should_gray_make_fade(rows: list[Row], row_index: int) -> bool:
    prev_bg = previous_background(rows, row_index)
    next_bg = find_next_content_background(rows, row_index + 1)
    if not prev_bg or not next_bg or prev_bg == next_bg:
        return False

    prev_row = next((row for row in reversed(rows[:row_index]) if not row.is_gray), None)
    next_row = next((row for row in rows[row_index + 1:] if not row.is_gray), None)
    if prev_row and next_row and (is_choice_row(prev_row) or is_follow_row(prev_row)) and (is_choice_row(next_row) or is_follow_row(next_row)):
        return False
    return True


def process_choice_group(builder: EpisodeBuilder, rows: list[Row], start: int) -> int:
    group_rows: list[tuple[Row, list[Row]]] = []
    i = start
    while i < len(rows):
        row = rows[i]
        if row.is_gray:
            i += 1
            continue
        if not is_choice_row(row):
            break

        follow_rows: list[Row] = []
        i += 1
        while i < len(rows) and is_follow_row(rows[i]):
            follow_rows.append(rows[i])
            i += 1
        group_rows.append((row, follow_rows))

    choice_node = {
        "type": "choice",
        "prompt": builder.last_prompt,
        "choices": [],
    }
    added = builder.add(choice_node)

    common_choice_bg = ""
    backgrounds = {row.background for row, _ in group_rows if row.background}
    if len(backgrounds) == 1:
        common_choice_bg = next(iter(backgrounds))

    branch_nodes: dict[str, int] = {}
    for choice_row, follow_rows in group_rows:
        choice = make_choice(choice_row)
        branch_target = branch_target_from_note(choice_row.note)
        for follow_row in follow_rows:
            choice["follow"].append(make_follow_line(follow_row))
            branch_target = branch_target or branch_target_from_note(follow_row.note)

        if branch_target:
            if branch_target not in branch_nodes:
                clear_node = builder.add({"type": "clear background"})
                move = {**EP3_BRANCH_MOVE, "after": branch_target}
                builder.add(move)
                branch_nodes[branch_target] = clear_node["id"]
            choice["nextNode"] = branch_nodes[branch_target]
        choice_node["choices"].append(choice)

    if not branch_nodes:
        next_id = builder.next_id
        if common_choice_bg and common_choice_bg != builder.current_background:
            transition = builder.pending_transition or make_transition(common_choice_bg)
            builder.pending_transition = None
            builder.ensure_background(common_choice_bg, transition)
        for choice in choice_node["choices"]:
            choice["nextNode"] = next_id

    added["choices"] = choice_node["choices"]
    return i


def process_rows(episode_id: str, rows: list[Row]) -> list[dict]:
    builder = EpisodeBuilder(episode_id)
    i = 0

    while i < len(rows):
        row = rows[i]
        if row.is_gray:
            if should_gray_make_fade(rows, i):
                next_bg = find_next_content_background(rows, i + 1)
                builder.pending_transition = make_transition(next_bg)
            i += 1
            continue

        if is_marker_row(row):
            i += 1
            continue

        if is_choice_row(row):
            i = process_choice_group(builder, rows, i)
            continue

        transition = None
        if row.background and row.background != builder.current_background:
            if builder.pending_transition:
                transition = builder.pending_transition
                builder.pending_transition = None
            elif note_has_fade(row):
                transition = make_transition(row.background)
            builder.ensure_background(row.background, transition)

        builder.add_row_characters(row)
        builder.add_sounds(row.sound)
        builder.add_dialogue(row)

        if note_has_fade(row) and (not row.background or row.background == builder.current_background):
            next_bg = find_next_content_background(rows, i + 1)
            if next_bg and next_bg != builder.current_background:
                builder.pending_transition = make_transition(next_bg)

        i += 1

    append_episode_end(builder)
    return builder.nodes


def append_episode_end(builder: EpisodeBuilder) -> None:
    if builder.episode_id in MINIGAME_MOVES:
        builder.add({"type": "clear background"})
        builder.add({**MINIGAME_MOVES[builder.episode_id]})
    elif builder.episode_id == "EP6 정기공연":
        builder.add({"type": "clear background"})
        builder.add({"type": "move", "next": "해피엔딩", "condition": {"affectionMin": 60}})
        builder.add({"type": "clear background"})
        builder.add({"type": "move", "next": "베드엔딩"})
    elif builder.episode_id == "해피엔딩":
        builder.add({"type": "clear background"})
        builder.add({"type": "dialogue", "speaker": "END", "text": "해피엔딩"})
    elif builder.episode_id == "베드엔딩":
        builder.add({"type": "clear background"})
        builder.add({"type": "dialogue", "speaker": "END", "text": "베드엔딩"})


def split_ending(rows: list[Row]) -> tuple[list[Row], list[Row]]:
    common: list[Row] = []
    happy: list[Row] = []
    bad: list[Row] = []
    target = common

    for row in rows:
        if row.selection == "<해피엔딩>":
            target = happy
            if row.background or row.sound or note_has_fade(row):
                marker_row = Row(**{**row.__dict__, "selection": "", "text": "", "speaker": ""})
                target.append(marker_row)
            continue
        if row.selection == "<베드엔딩>":
            target = bad
            if row.background or row.sound or note_has_fade(row):
                marker_row = Row(**{**row.__dict__, "selection": "", "text": "", "speaker": ""})
                target.append(marker_row)
            continue
        target.append(row)

    return common + happy, common + bad


def generate() -> None:
    workbook = load_workbook(SOURCE_XLSX, data_only=False)
    sheets = {ws.title.strip(): ws for ws in workbook.worksheets}
    episodes: dict[str, list[dict]] = {}

    for sheet_name in STORY_SHEETS:
        episodes[sheet_name] = process_rows(sheet_name, read_rows(sheets[sheet_name]))

    happy_rows, bad_rows = split_ending(read_rows(sheets["엔딩"]))
    episodes["해피엔딩"] = process_rows("해피엔딩", happy_rows)
    episodes["베드엔딩"] = process_rows("베드엔딩", bad_rows)

    js = (
        "// Generated from ../도파민때문에_시나리오_v1.0.xlsx.\n"
        "// Run scripts/generate_story_v1.py from the project root to regenerate.\n"
        'const STORY_START_EPISODE = "EP1 첫만남";\n'
        "const EPISODES = "
        + json.dumps(episodes, ensure_ascii=False, indent=2)
        + ";\n"
    )
    OUTPUT_JS.write_text(js, encoding="utf-8")


if __name__ == "__main__":
    generate()
